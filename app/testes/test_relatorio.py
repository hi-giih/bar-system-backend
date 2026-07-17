import io

from openpyxl import load_workbook


def _adiciona_produto(client, auth_headers, comanda_id, produto_id, quantidade):
    return client.post(
        f"/comanda/{comanda_id}/produtos",
        json={"produto_id": produto_id, "quantidade": quantidade},
        headers=auth_headers,
    )


def test_baixar_planilha_sem_autenticacao_bloqueado(client):
    r = client.get("/relatorios/comandas")
    assert r.status_code == 401


def test_baixar_planilha_retorna_xlsx_valido(client, auth_headers, comanda_id, produto_id):
    _adiciona_produto(client, auth_headers, comanda_id, produto_id, 2)  # total = 20

    r = client.get("/relatorios/comandas", headers=auth_headers)
    assert r.status_code == 200
    assert r.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in r.headers["Content-Disposition"]

    workbook = load_workbook(io.BytesIO(r.data))
    assert workbook.sheetnames == ["Comandas Abertas", "Histórico"]


def test_planilha_lista_comanda_aberta_com_produtos(client, auth_headers, comanda_id, produto_id):
    _adiciona_produto(client, auth_headers, comanda_id, produto_id, 3)  # total = 30

    r = client.get("/relatorios/comandas", headers=auth_headers)
    workbook = load_workbook(io.BytesIO(r.data))
    aba = workbook["Comandas Abertas"]

    linhas = list(aba.iter_rows(values_only=True))
    assert linhas[0] == (
        "Cliente", "Data", "Produto", "Quantidade", "Preço Unitário", "Subtotal", "Saldo Devedor",
    )
    assert linhas[1] == ("Cliente Teste", "2026-07-06", "Produto Teste", 3, 10.0, 30.0, 30.0)


def test_planilha_lista_comanda_fechada_no_historico(client, auth_headers, comanda_id, produto_id):
    _adiciona_produto(client, auth_headers, comanda_id, produto_id, 1)  # total = 10
    client.post(
        "/pagamento/",
        json={"comanda_id": comanda_id, "valor": 10, "forma_pagamento": "dinheiro"},
        headers=auth_headers,
    )
    client.post(f"/comanda/{comanda_id}/fechar", headers=auth_headers)

    r = client.get("/relatorios/comandas", headers=auth_headers)
    workbook = load_workbook(io.BytesIO(r.data))

    abertas = list(workbook["Comandas Abertas"].iter_rows(values_only=True))
    assert len(abertas) == 1  # so o cabecalho, a comanda fechou

    historico = list(workbook["Histórico"].iter_rows(values_only=True))
    assert historico[1] == (
        "Cliente Teste", "2026-07-06", "dinheiro", "Produto Teste", 1, 10.0, 10.0, 10.0,
    )
